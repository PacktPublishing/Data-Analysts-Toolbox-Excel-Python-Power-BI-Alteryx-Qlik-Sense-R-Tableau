#! /usr/bin/python3
# Add shebang line above

# Import modules
try:
    import os
    import csv
    from openpyxl import load_workbook
except ModuleNotFoundError as error:
    print(f'There was a problem importing a module: {error}')
    raise

# Create functions to clean up code
def next_available_row(col):
    for row in range(5, ws.max_row+1):
        if ws.cell(row, col).value is None:
            return row
            break

def insert_transactions(row_start, transactions, col_start, col_end):
    if row_start is None:
        raise Exception('No empty row found in spreadsheet.')
    else:
        for row in enumerate(range(row_start, row_start + len(transactions))):
            for col in enumerate(range(col_start, col_end)):
                ws.cell(row[1], col[1], value=transactions[row[0]][col[0]])

par_dir_path = os.path.dirname(os.path.abspath(__file__))

income = []
expenses = []

# Open transactions.csv
with open(os.path.join(par_dir_path, 'transactions.csv'), newline='') as csvfile:
    transactions = csv.DictReader(csvfile, delimiter=',')
    rows = list(transactions)
    for row in rows:
        # Separate expenses from income
        if float(row['Amount'].replace(',', '')) > 0:
            income.append([
                row['Date'],
                float(row['Amount'].replace(',', '')),
                row['Simple Description'],
                row['Category']
                ])
        else:
            expenses.append([
                row['Date'],
                float(row['Amount'].replace(',', '')),
                row['Simple Description'],
                row['Category']
                ])

# Load Excel workbook
try:
    wb = load_workbook(filename=os.path.join(par_dir_path, 'budget.xlsx'))
except:
    print('Double check that the file is in the right folder with correct ext.')
    raise

try:
    ws = wb['Transactions']

    # Find next available expense row
    exp_starting_row = next_available_row(3)

    # Find next available income row
    inc_starting_row = next_available_row(8)

    # Insert transactions
    insert_transactions(exp_starting_row, expenses, 2, 6)
    insert_transactions(inc_starting_row, income, 7, 11)

    ws.append([''])
finally:
    # Save workbook
    wb.save(os.path.join(par_dir_path, 'budget.xlsx'))
